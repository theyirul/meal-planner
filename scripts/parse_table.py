#!/usr/bin/env python3
"""
매일아침 — 통합 테이블 파서

영등포구, 성남시, 동대문구, 용인시 등 "큰 테이블" 형태의 식단표를 하나의 파서로 처리.
포맷별 차이(열 위치, 알레르기 표기, 페이지)는 자동 감지.

지원 포맷:
  - yeongdeungpo: 19열, 1페이지, 원형숫자 알레르기
  - seongnam: 11열, 1페이지, 인라인 콤마 알레르기
  - dongdaemun: 16열, 2페이지, 원형숫자 알레르기
  - yongin: 가변열, 1페이지, 원형숫자 알레르기, 요일열 동적 감지
  - gangseo: 11열, 1페이지, 원형숫자 알레르기, 한글 주차 표기

사용법:
  from parse_table import parse_table_pdf, parse_recipe_xlsx, parse_recipe_pdf
"""

import re
from datetime import date
from pathlib import Path

import pdfplumber

from parse_common import (
    extract_allergy_circled,
    extract_allergy_inline,
    clean_menu_name,
    build_output_json,
    DECO_CHARS,
    CIRCLED_PATTERN,
)


# ── 포맷 설정 ──

FORMAT_CONFIGS = {
    "yeongdeungpo": {
        "page": 0,
        "allergy_fn": extract_allergy_circled,
        "day_cols": [1, 5, 8, 11, 13, 16],  # 월~토 in 19-col
        "week_detect": "empty_first",  # col[0] 비어있으면 날짜행
        "meal_offsets": {"오전간식": 1, "점심": 2, "오후간식": 3},
        "block_size": 5,
    },
    "seongnam": {
        "page": 0,
        "allergy_fn": extract_allergy_inline,
        "day_cols": [3, 5, 7, 8, 9, 10],  # 월~토 in 11-col
        "week_detect": "week_label",  # "N주차" 텍스트
        "meal_offsets": {"오전간식": 1, "점심": 2, "오후간식": 3},
        "block_size": 6,
        "merge_prefix": r'경기도과일\s*또는\s*\n',
    },
    "dongdaemun": {
        "page": 1,
        "allergy_fn": extract_allergy_circled,
        "day_cols": [1, 4, 6, 8, 11, 14],  # 월~토 in 16-col
        "week_detect": "date_label",  # "일 자" 텍스트
        "meal_offsets": {"오전간식": 1, "점심": 2, "오후간식": 3},
        "block_size": 5,
    },
    "yongin": {
        "page": 0,
        "allergy_fn": extract_allergy_circled,
        "day_cols": "auto",  # 헤더 행에서 동적 감지 (월/화/수/목/금/토)
        "week_detect": "date_keyword",  # col[0]에 '날짜' 텍스트
        "meal_offsets": {"오전간식": 1, "점심": 2, "오후간식": 3},
        "block_size": 6,  # 날짜 + 오전간식 + 점심 + 오후간식 + 열량 + 원산지
    },
    "gangseo": {
        "page": 0,
        "allergy_fn": extract_allergy_circled,
        "day_cols": "auto_per_week",  # 주마다 일자 행에서 동적 감지 (월별 열 수 가변)
        "week_detect": "week_korean",  # col[0]에 '첫째주', '둘째주' 등
        "meal_offsets": {"오전간식": 1, "점심": 2, "오후간식": 3},
        "block_size": 5,
    },
}


def _has_date_keyword_rows(table: list) -> bool:
    """테이블 내에 '날짜' 키워드 행이 있으면 True (용인시 감지용)"""
    for row in table:
        if row and '날짜' in str(row[0] or '').strip():
            return True
    return False


def _detect_day_cols_from_header(table: list) -> list[int]:
    """
    헤더 행에서 요일(월/화/수/목/금/토) 열 위치를 동적으로 감지.
    용인시처럼 월별로 열 수가 달라지는 경우에 사용.
    """
    day_names = ['월', '화', '수', '목', '금', '토']
    result = []
    # 처음 5행 안에서 요일 헤더 찾기
    for row in table[:5]:
        cols = []
        for ci, cell in enumerate(row):
            val = str(cell or '').strip()
            if val in day_names:
                cols.append(ci)
        if len(cols) >= 5:  # 최소 월~금 5개
            return cols
    return result


def detect_table_format(pdf_path: str) -> str | None:
    """PDF를 열어서 테이블 포맷을 자동 감지. None이면 테이블 포맷 아님."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # 1페이지 먼저 확인
            tables = pdf.pages[0].extract_tables()
            if tables:
                biggest = max(tables, key=lambda t: len(t))
                ncols = len(biggest[0]) if biggest else 0

                # 강서구: col[0]에 한글 주차('째주') + col[1]에 '일자'/'요일' (열 수 가변: 11~17+)
                first = str(biggest[0][0] or '').strip()
                second = str(biggest[0][1] or '').strip() if len(biggest[0]) > 1 else ''
                if re.search(r'째\s*주|째\n주', first) and ('일자' in second or '요일' in second):
                    return "gangseo"

                # 용인시: 요일 헤더 + '날짜' 키워드 행 (열 수 가변)
                if ncols >= 12:
                    header_days = [str(c or '') for c in biggest[0]]
                    if '월' in header_days and '화' in header_days:
                        if _has_date_keyword_rows(biggest):
                            return "yongin"
                        # 영등포구: 요일 헤더 있지만 '날짜' 행 없음
                        if ncols >= 15:
                            return "yeongdeungpo"

                # 성남시: 11열, "N주차"
                if ncols == 11:
                    first = str(biggest[0][0] or '').strip()
                    if re.match(r'\d+주차', first):
                        return "seongnam"

            # 2페이지 확인 (동대문구 등)
            if len(pdf.pages) >= 2:
                tables2 = pdf.pages[1].extract_tables()
                if tables2:
                    biggest2 = max(tables2, key=lambda t: len(t))
                    ncols2 = len(biggest2[0]) if biggest2 else 0
                    if ncols2 >= 14:
                        first = str(biggest2[0][0] or '').strip()
                        if '일' in first and '자' in first:
                            return "dongdaemun"

        return None
    except Exception:
        return None


def _detect_year_month(pdf, fmt_config: dict) -> tuple[int, int]:
    """PDF 텍스트에서 연도/월 감지"""
    page_idx = fmt_config.get("page", 0)
    text = ''
    for p in pdf.pages[:page_idx + 1]:
        text += (p.extract_text() or '') + '\n'
    # null 문자 및 특수 공백 제거 (일부 PDF에서 \x00 포함)
    text = text.replace('\x00', ' ')

    # "(N월 식단)" 패턴
    m = re.search(r'\(?(\d{1,2})월\s*식단\)?', text)
    if m:
        month = int(m.group(1))
        ym = re.search(r'(\d{4})\s*[년.]', text)
        year = int(ym.group(1)) if ym else date.today().year
        return year, month

    # "2026년 3월" 패턴
    m = re.search(r'(\d{4})\s*[년.]\s*(\d{1,2})\s*월', text)
    if m:
        return int(m.group(1)), int(m.group(2))

    # "3. 일반식" + "[2026년3월" 패턴
    m = re.search(r'(\d{4})년\s*(\d{1,2})월', text)
    if m:
        return int(m.group(1)), int(m.group(2))

    # 발행일에서 추론
    m = re.search(r'(\d{4})\.\s*(\d{1,2})\.\s*\d{1,2}', text)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        return (y, mo + 1) if mo < 12 else (y + 1, 1)

    raise ValueError("연도/월을 감지할 수 없습니다")


def _extract_date_from_cell(cell_text: str) -> int | None:
    """셀에서 날짜 숫자 추출. "9(월)푸드브릿지Day" → 9"""
    if not cell_text:
        return None
    m = re.match(r'(\d{1,2})', str(cell_text).strip())
    return int(m.group(1)) if m else None


def _is_date_row(row: list, fmt: str, config: dict) -> bool:
    """이 행이 날짜 행인지 판별"""
    method = config["week_detect"]
    first = str(row[0] or '').strip()

    if method == "empty_first":
        # 영등포구: col[0]이 비어있고 다른 열에 숫자가 있으면 날짜행
        if first and first not in ('', ' '):
            return False
        for ci in config["day_cols"]:
            if ci < len(row):
                dn = _extract_date_from_cell(str(row[ci] or ''))
                if dn and 1 <= dn <= 31:
                    return True
        return False

    elif method == "week_label":
        # 성남시: "N주차" 텍스트
        return bool(re.match(r'\d+주차', first))

    elif method == "date_label":
        # 동대문구: "일 자" 또는 "일자" 텍스트
        return '일' in first and '자' in first

    elif method == "week_korean":
        # 강서구: col[0]에 한글 주차 표기 ('첫째주', '둘째주' 등) 또는 col[1]에 '일자'/'요일'
        if first and re.search(r'째\s*주|째\n주', first):
            return True
        second = str(row[1] or '').strip() if len(row) > 1 else ''
        if '일자' in second or '요일' in second:
            return True
        return False

    elif method == "date_keyword":
        # 용인시: col[0]에 '날짜' 텍스트
        if '날짜' in first:
            return True
        # 마지막 주 fallback: col[0]이 정확히 빈 문자열 '' (None 아님)
        # 주석/푸터 행은 col[0]이 None이므로 구분 가능
        if row[0] == '':
            count = sum(
                1 for ci in range(1, min(len(row), 14))
                if _extract_date_from_cell(str(row[ci] or ''))
            )
            return count >= 1
        return False

    return False


def _parse_menu_cell(cell_text: str, allergy_fn, merge_prefix: str | None = None) -> list[dict]:
    """한 셀에서 메뉴 아이템들 추출"""
    if not cell_text or not cell_text.strip():
        return []

    items = []

    # 접두어 합치기 (경기도과일 또는)
    if merge_prefix:
        cell_text = re.sub(merge_prefix, lambda m: m.group().replace('\n', ' '), cell_text)

    # 전처리: 괄호 안 보조설명 제거 (영등포구 스타일)
    cell_text = re.sub(r'\n?\(([^)]*[/][^)]*)\)', '', cell_text)
    # 대괄호 처리:
    # - 대괄호 안 줄바꿈 제거 (병합 셀에서 "[메뉴명\n①②③]" 형태 발생)
    # - 알레르기 번호 포함: "[오곡밥①②⑤⑥]" → "(오곡밥①②⑤⑥)" (대체메뉴로 처리)
    # - 숫자 없음: "[오렌지]" → "" (대체메뉴 표기지만 생략)
    cell_text = re.sub(r'\[([^\]]+)\]', lambda m: '[' + m.group(1).replace('\n', '') + ']', cell_text)
    def _bracket_replace(m):
        inner = m.group(1)
        if re.search(r'[①-⑲]|\d', inner):
            return f'({inner})'
        return ''
    cell_text = re.sub(r'\[([^\]]+)\]', _bracket_replace, cell_text)

    lines = cell_text.strip().split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        if not line:
            continue

        # 순수 숫자 줄 → 이전 아이템의 알레르기
        if re.match(r'^[\d,\s]+$', line) and items:
            for n_str in re.findall(r'\d{1,2}', line):
                v = int(n_str)
                if 1 <= v <= 19 and v not in items[-1]['allergy']:
                    items[-1]['allergy'].append(v)
            items[-1]['allergy'] = sorted(set(items[-1]['allergy']))
            continue

        # 콤마로 시작하는 줄 → 이전 알레르기 연속
        if re.match(r'^,\s*\d', line) and items:
            for n_str in re.findall(r'\d{1,2}', line):
                v = int(n_str)
                if 1 <= v <= 19 and v not in items[-1]['allergy']:
                    items[-1]['allergy'].append(v)
            items[-1]['allergy'] = sorted(set(items[-1]['allergy']))
            continue

        # 열량 정보 스킵
        if re.match(r'^\d+/\d+', line):
            continue

        # 라벨 스킵
        if line in ('(공통)', '점심', '오전간식', '오후간식', '오후 간식', '(숭늉)'):
            continue
        # 안내 문구 스킵 (공지사항, 가정통신문 등)
        if re.search(r'[※★☆]|가정통신문|확인해\s*주세요|자세한\s*내용', line):
            continue

        # / 구분자로 분리 (예: "♣찐감자/우유②")
        sub_parts = _split_by_separator(line)
        for part in sub_parts:
            part = part.strip()
            if not part:
                continue
            # 후식: 접두어 제거
            part = re.sub(r'^후식\s*:\s*', '', part)
            # 괄호 안 대체메뉴: "(돈육무조림⑤⑥⑩)" → 별도 아이템으로
            # 단, "(+✔부럼④⑭)" 같은 추가메뉴도 처리
            alt_items = _extract_alt_items(part)
            for alt in alt_items:
                name, allergy = allergy_fn(alt)
                name = clean_menu_name(name)
                if name:
                    items.append({"name": name, "allergy": allergy})

    return items


def _split_by_separator(text: str) -> list[str]:
    """괄호 안의 /를 무시하고 /로 분리"""
    parts = []
    depth = 0
    current = ''
    for ch in text:
        if ch == '(':
            depth += 1
            current += ch
        elif ch == ')':
            depth = max(0, depth - 1)
            current += ch
        elif ch == '/' and depth == 0:
            if current.strip():
                parts.append(current.strip())
            current = ''
        else:
            current += ch
    if current.strip():
        parts.append(current.strip())

    # 순수 숫자 파트는 이전 파트에 합치기
    merged = []
    for p in parts:
        if re.match(r'^[\d\s]+$', p) and merged:
            merged[-1] = merged[-1] + '/' + p
        else:
            merged.append(p)
    return merged if merged else [text]


def _extract_alt_items(text: str) -> list[str]:
    """
    괄호 안 대체메뉴를 별도 분리.
    "고등어무조림⑤⑥⑦\n(돈육무조림⑤⑥⑩)" → ["고등어무조림⑤⑥⑦", "돈육무조림⑤⑥⑩"]
    "(+✔부럼④⑭)" → ["+✔부럼④⑭"]
    "백미밥(오곡밥)" → ["백미밥(오곡밥)"] (알레르기 번호 없으면 그대로)
    """
    # 대체메뉴 괄호: 숫자나 원형숫자가 포함된 독립 괄호
    pattern = re.compile(r'\((\+?[^)]*(?:[①-⑲]|\d{1,2}(?:,\d{1,2})*)[^)]*)\)')
    alts = []
    remaining = text
    for m in reversed(list(pattern.finditer(text))):
        inner = m.group(1)
        # 연령 표기 스킵
        if re.match(r'만\s*\d', inner):
            continue
        # 대체메뉴 추출
        alts.insert(0, inner)
        remaining = remaining[:m.start()] + remaining[m.end():]

    result = []
    if remaining.strip():
        result.append(remaining.strip())
    result.extend(alts)
    return result if result else [text]


def parse_table_pdf(pdf_path: str, fmt: str | None = None) -> dict:
    """
    테이블 기반 식단표 PDF 파싱.

    Args:
        pdf_path: PDF 파일 경로
        fmt: 포맷 이름 (None이면 자동 감지)

    Returns:
        {"year": int, "month": int, "menus": {date_str: [items]}}
    """
    if fmt is None:
        fmt = detect_table_format(pdf_path)
    if fmt not in FORMAT_CONFIGS:
        raise ValueError(f"알 수 없는 포맷: {fmt}")

    config = FORMAT_CONFIGS[fmt]
    allergy_fn = config["allergy_fn"]
    merge_prefix = config.get("merge_prefix")

    with pdfplumber.open(pdf_path) as pdf:
        year, month = _detect_year_month(pdf, config)
        page_idx = config["page"]
        tables = pdf.pages[page_idx].extract_tables()

    if not tables:
        raise ValueError("테이블을 찾을 수 없습니다")

    table = max(tables, key=lambda t: len(t))
    ncols = len(table[0]) if table else 0
    print(f"  [{fmt}] 감지: {year}년 {month}월, {len(table)}행 x {ncols}열")

    # 동적 day_cols 감지 (용인시 등 가변 열 포맷)
    day_cols_raw = config["day_cols"]
    use_merged_cells = (day_cols_raw == "auto")
    auto_per_week = (day_cols_raw == "auto_per_week")
    if use_merged_cells:
        day_cols = _detect_day_cols_from_header(table)
        if not day_cols:
            raise ValueError("요일 헤더를 감지할 수 없습니다")
    elif auto_per_week:
        day_cols = []  # 주마다 동적으로 결정
    else:
        day_cols = day_cols_raw

    # 셀 병합 함수: start_col~end_col 범위의 non-None 값 합치기
    def _get_merged_cell(row: list, start_col: int, end_col: int) -> str:
        parts = []
        for ci in range(start_col, min(end_col, len(row))):
            val = str(row[ci] or '').strip()
            if val:
                parts.append(val)
        return ''.join(parts)

    menus = {}
    row_idx = 0

    while row_idx < len(table):
        row = table[row_idx]

        if not _is_date_row(row, fmt, config):
            row_idx += 1
            continue

        # auto_per_week: 이 주의 일자 행에서 날짜가 있는 열을 동적 감지
        week_day_cols = day_cols
        if auto_per_week:
            week_day_cols = []
            for ci in range(len(row)):
                dn = _extract_date_from_cell(str(row[ci] or ''))
                if dn and 1 <= dn <= 31:
                    week_day_cols.append(ci)

        # 날짜 열에서 날짜 추출
        date_cols = {}
        for ci in week_day_cols:
            if ci >= len(row):
                continue
            dn = _extract_date_from_cell(str(row[ci] or ''))
            if dn and 1 <= dn <= 31:
                date_cols[ci] = dn

        if not date_cols:
            row_idx += 1
            continue

        # 각 날짜에서 끼니별 메뉴 추출
        sorted_day_cols = sorted(date_cols.keys())
        for i, ci in enumerate(sorted_day_cols):
            dn = date_cols[ci]
            try:
                ds = date(year, month, dn).isoformat()
            except ValueError:
                continue

            # 다음 날짜 열 위치 (셀 병합 범위 계산용)
            next_ci = sorted_day_cols[i + 1] if i + 1 < len(sorted_day_cols) else len(row)

            day_items = []
            for sec_name, offset in config["meal_offsets"].items():
                ri = row_idx + offset
                if ri >= len(table):
                    break
                sec_row = table[ri]
                if ci >= len(sec_row):
                    continue
                # 가변 열 포맷(용인시)은 병합 읽기, 고정 열 포맷은 단일 셀 읽기
                if use_merged_cells:
                    cell_text = _get_merged_cell(sec_row, ci, next_ci)
                else:
                    cell_text = str(sec_row[ci] or '')
                for item in _parse_menu_cell(cell_text, allergy_fn, merge_prefix):
                    item['sec'] = sec_name
                    day_items.append(item)

            if day_items:
                menus[ds] = day_items

        row_idx += config["block_size"]

    print(f"  [{fmt}] {len(menus)}일 추출, 총 {sum(len(v) for v in menus.values())}개 아이템")
    return {"year": year, "month": month, "menus": menus}


def parse_recipe_xlsx(xlsx_path: str) -> dict:
    """
    레시피 Excel 파싱 (영등포구/성남시 공통 — 주별 시트 구조).

    Returns:
        {"메뉴명": "조리방법 텍스트", ...}
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    recipes = {}

    # 주별 시트 또는 월별 시트 찾기
    # 패턴: "N주차", "N월", "M.D~M.D" (용인시 스타일) 등
    target_sheets = [s for s in wb.sheetnames if re.search(r'\d+주|\d+월|\d+\.\d+', s)]
    # 아무 시트도 안 걸리면 전체 시트 처리
    if not target_sheets:
        target_sheets = wb.sheetnames

    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        # 헤더 찾기: "음식명"과 "조리방법"
        header_row = None
        col_food = None
        col_recipe = None

        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or '').strip()
                if val == '음식명':
                    header_row = r
                    col_food = c
                elif val == '조리방법':
                    col_recipe = c
            if header_row and col_food and col_recipe:
                break

        if not header_row or not col_food or not col_recipe:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            food_val = ws.cell(r, col_food).value
            recipe_val = ws.cell(r, col_recipe).value

            if not food_val:
                continue
            food_name = str(food_val).strip()
            food_name = food_name.replace('_x000D_', '').strip()
            food_name = re.sub(r'\(만\s*\d+-?\d*세\)\s*\n?', '', food_name).strip()
            food_name = DECO_CHARS.sub('', food_name).strip()
            if not food_name:
                continue

            recipe_text = str(recipe_val).strip() if recipe_val else ''
            recipe_text = recipe_text.replace('_x000D_', '')

            if recipe_text and recipe_text not in ('None', '-', ''):
                if food_name not in recipes:
                    recipes[food_name] = recipe_text

    print(f"  [Excel 레시피] {len(recipes)}개 추출")
    return recipes


def parse_recipe_pdf(pdf_path: str) -> dict:
    """
    레시피 PDF 파싱 (동대문구 등 — 6열 테이블 구조).

    Returns:
        {"메뉴명": "조리방법 텍스트", ...}
    """
    recipes = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 6:
                        continue
                    food_name = str(row[2] or '').strip()
                    recipe_text = str(row[5] or '').strip()

                    if not food_name or food_name in ('음식명', ''):
                        continue

                    # 알레르기 표기 제거: "소고기채소덮밥 ⑤⑥⑯" → "소고기채소덮밥"
                    clean_name, _ = extract_allergy_circled(food_name)
                    clean_name = clean_menu_name(clean_name)
                    # 괄호 표기도 제거: "(부럼 ④⑭)" → 스킵
                    if clean_name.startswith('(') and clean_name.endswith(')'):
                        continue
                    clean_name = clean_name.replace('_x000D_', '').strip()

                    if not clean_name:
                        continue

                    recipe_text = recipe_text.replace('_x000D_', '')

                    if recipe_text and recipe_text not in ('None', '-', ''):
                        if clean_name not in recipes:
                            recipes[clean_name] = recipe_text

    print(f"  [PDF 레시피] {len(recipes)}개 추출")
    return recipes


# ── CLI ──
if __name__ == '__main__':
    import json
    import sys

    if len(sys.argv) < 2:
        print("Usage: python3 parse_table.py <식단표.pdf> [레시피.xlsx|pdf] [output.json]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    recipe_path = sys.argv[2] if len(sys.argv) >= 3 and not sys.argv[2].endswith('.json') else None
    out_path = sys.argv[-1] if sys.argv[-1].endswith('.json') else Path(pdf_path).stem + '.json'

    fmt = detect_table_format(pdf_path)
    print(f"포맷 감지: {fmt}")

    pdf_data = parse_table_pdf(pdf_path, fmt)

    recipes = {}
    if recipe_path:
        if recipe_path.endswith('.xlsx'):
            recipes = parse_recipe_xlsx(recipe_path)
        elif recipe_path.endswith('.pdf'):
            recipes = parse_recipe_pdf(recipe_path)

    data = build_output_json(pdf_data['year'], pdf_data['month'], pdf_data['menus'], recipes)

    Path(out_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n[OK] → {out_path}")
    print(f"  {data['year']}년 {data['month']}월, {len(data['menus'])}일, "
          f"{sum(len(v) for v in data['menus'].values())}개 아이템")
