#!/usr/bin/env python3
"""
매일아침 — 성남시 포맷 파서

성남시어린이급식관리지원센터 식단표 PDF + 레시피 Excel 파서.
다른 포맷과의 차이점:
  - PDF: 34행 x 11열 단일 테이블, 인라인 콤마형 알레르기 (예: 5,6,18)
  - Excel: 주별 시트 ("4월 1주(1일~4일)"), 7열 구조
  - 주마다 6행 블록: 주차명(날짜), 오전간식, 점심, 오후간식, 열량/단백질, 원산지

사용법:
  from parse_seongnam import parse_sn_pdf, parse_sn_excel, build_sn_json
"""

import re
from datetime import date, timedelta
from pathlib import Path

import pdfplumber

# ── 19종 알레르기 ──
ALLERGEN_NAMES = {
    1: "난류(계란)", 2: "우유", 3: "메밀", 4: "땅콩", 5: "대두",
    6: "밀", 7: "고등어", 8: "게", 9: "새우", 10: "돼지고기",
    11: "복숭아", 12: "토마토", 13: "아황산류", 14: "호두", 15: "닭고기",
    16: "쇠고기", 17: "오징어", 18: "조개류", 19: "잣"
}

# 메뉴명에서 제거할 장식 문자
DECO_CHARS = re.compile(r'[♬★◉♥\U000f0074]')

# 날짜 열 인덱스 (11열 테이블)
# 주 1: 시작 요일에 따라 가변, 보통 col 7~10
# 주 2+: col 3(월), 5(화), 7(수), 8(목), 9(금), 10(토)
ALL_DAY_COLS = [3, 5, 7, 8, 9, 10]


def _extract_allergy_inline(text: str) -> tuple[str, list[int]]:
    """
    인라인 콤마형 알레르기 번호 추출.
    "조갯살호박국5,6,18" → ("조갯살호박국", [5, 6, 18])
    "미니돈가스&소스" + 다음줄 "1,5,6,10,12,16" 은 호출 전에 합쳐서 전달.
    "백김치9(백깍두기9)" → ("백김치", [9]) + 대체품 별도 처리
    """
    if not text or not text.strip():
        return ('', [])

    text = text.strip()
    nums = []

    # 괄호 안 대체메뉴 처리: "백김치9(백깍두기9)" → 대체메뉴 제거
    # 단, "(만1-2세)" 같은 연령 표기는 유지 후 별도 제거
    alt_pattern = re.compile(r'\(([^)]*\d+[^)]*)\)')
    alt_matches = list(alt_pattern.finditer(text))
    for m in reversed(alt_matches):
        inner = m.group(1)
        # 연령 표기 스킵
        if re.match(r'만\s*\d', inner):
            continue
        # 대체메뉴 괄호 전체 제거
        text = text[:m.start()] + text[m.end():]

    # 원형숫자 처리 (혼용 대비)
    circled = {'①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
               '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10,
               '⑪': 11, '⑫': 12, '⑬': 13, '⑭': 14, '⑮': 15,
               '⑯': 16, '⑰': 17, '⑱': 18, '⑲': 19}
    for ch, n in circled.items():
        if ch in text:
            nums.append(n)
            text = text.replace(ch, '')

    # 끝에 붙은 숫자+콤마 패턴: "조갯살호박국5,6,18" or "백김치9"
    # 패턴: 메뉴명 뒤에 숫자,숫자,... (공백 없이 또는 약간의 공백)
    # 가장 뒤에서부터 숫자+콤마 시퀀스를 찾기
    m = re.search(r'[\s,]*(\d{1,2}(?:\s*,\s*\d{1,2})*)\s*$', text)
    if m:
        num_str = m.group(1)
        for n_str in re.findall(r'\d{1,2}', num_str):
            v = int(n_str)
            if 1 <= v <= 19 and v not in nums:
                nums.append(v)
        text = text[:m.start()].strip()

    # 연령 표기 제거
    text = re.sub(r'\(만\s*\d+-?\d*세\)', '', text).strip()
    # 장식 문자 제거
    text = DECO_CHARS.sub('', text).strip()

    return (text, sorted(set(nums)))


def _parse_menu_cell(cell_text: str) -> list[dict]:
    """
    한 셀의 텍스트에서 개별 메뉴 아이템을 추출.
    줄바꿈으로 구분, 각 아이템에서 인라인 알레르기 분리.
    다음 줄이 순수 숫자이면 이전 아이템의 알레르기로 합침.
    """
    if not cell_text or not cell_text.strip():
        return []

    items = []
    lines = cell_text.strip().split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        i += 1
        if not line:
            continue

        # 순수 숫자 줄 체크 (이전 아이템의 알레르기)
        if re.match(r'^[\d,\s]+$', line) and items:
            for n_str in re.findall(r'\d{1,2}', line):
                v = int(n_str)
                if 1 <= v <= 19 and v not in items[-1]['allergy']:
                    items[-1]['allergy'].append(v)
            items[-1]['allergy'] = sorted(set(items[-1]['allergy']))
            continue

        # 콤마로 시작하는 줄도 이전 아이템 알레르기 연속
        if re.match(r'^,\s*\d', line) and items:
            for n_str in re.findall(r'\d{1,2}', line):
                v = int(n_str)
                if 1 <= v <= 19 and v not in items[-1]['allergy']:
                    items[-1]['allergy'].append(v)
            items[-1]['allergy'] = sorted(set(items[-1]['allergy']))
            continue

        # 열량/영양 정보 스킵
        if re.match(r'^\d+/\d+$', line):
            continue

        # 라벨 스킵
        if line in ('(공통)', '점심', '오전간식', '오후간식', '오후 간식'):
            continue

        # 콤마로 구분된 간식 분리: "빵(머핀1,2,5,6),우유2"
        # 이건 좀 복잡 — 콤마가 알레르기 번호와 메뉴 구분 양쪽에 사용됨
        # 간식 줄에서 "),..." 패턴으로 분리
        sub_items = _split_snack_items(line)
        for sub in sub_items:
            name, allergy = _extract_allergy_inline(sub)
            if name:
                items.append({"name": name, "allergy": allergy})

    return items


def _split_snack_items(text: str) -> list[str]:
    """
    간식 셀에서 쉼표로 구분된 개별 아이템 분리.
    "빵(머핀1,2,5,6),우유2" → ["빵(머핀1,2,5,6)", "우유2"]
    "찐(군)감자,우유2" → ["찐(군)감자", "우유2"]
    핵심: 괄호 안의 콤마는 무시하고, 괄호 밖의 콤마만 분리.
    """
    parts = []
    depth = 0
    current = ''
    for ch in text:
        if ch == '(':
            depth += 1
            current += ch
        elif ch == ')':
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0:
            # 괄호 밖 콤마 → 분리 후보
            # 하지만 숫자만 이어지면 알레르기 번호이므로 분리하지 않음
            rest = text[text.index(current) + len(current) + 1:] if current in text else ''
            # 간단 판별: current 뒤의 콤마 다음이 한글로 시작하면 메뉴 분리
            trimmed = current.strip()
            if trimmed:
                parts.append(trimmed)
            current = ''
        else:
            current += ch
    if current.strip():
        parts.append(current.strip())

    # parts가 분리되었는데, 순수 숫자 파트가 있으면 이전 파트에 합치기
    merged = []
    for p in parts:
        if re.match(r'^[\d\s]+$', p) and merged:
            merged[-1] = merged[-1] + ',' + p
        else:
            merged.append(p)

    return merged if merged else [text]


def _extract_date_from_header(cell_text: str) -> int | None:
    """주차 행의 날짜 셀에서 숫자 추출. "1(수)", "6(월)", "♥생일식단♥" 등"""
    if not cell_text:
        return None
    cell_text = str(cell_text).strip()
    m = re.match(r'(\d{1,2})\s*[\(\[（]', cell_text)
    if m:
        return int(m.group(1))
    # 순수 숫자
    m = re.match(r'^(\d{1,2})$', cell_text)
    if m:
        return int(m.group(1))
    return None


def _detect_year_month(pdf) -> tuple[int, int]:
    """PDF 텍스트에서 연도/월 감지"""
    text = pdf.pages[0].extract_text() or ''

    # "4월 식단" 패턴
    m = re.search(r'(\d{1,2})월\s*식단', text)
    if m:
        month = int(m.group(1))
        # 연도: "2026년" 패턴
        ym = re.search(r'(\d{4})\s*[년.]', text)
        year = int(ym.group(1)) if ym else date.today().year
        return year, month

    # "2026년 4월" 패턴
    m = re.search(r'(\d{4})\s*[년.]\s*(\d{1,2})\s*월', text)
    if m:
        return int(m.group(1)), int(m.group(2))

    raise ValueError("연도/월을 감지할 수 없습니다")


def parse_sn_pdf(pdf_path: str) -> dict:
    """
    성남시 식단표 PDF 파싱.

    Returns:
        {
            "year": 2026, "month": 4,
            "menus": {
                "2026-04-01": [
                    {"name": "백미밥", "allergy": [], "sec": "점심"},
                    ...
                ], ...
            }
        }
    """
    with pdfplumber.open(pdf_path) as pdf:
        year, month = _detect_year_month(pdf)
        tables = pdf.pages[0].extract_tables()

    if not tables:
        raise ValueError("테이블을 찾을 수 없습니다")

    table = max(tables, key=lambda t: len(t))
    ncols = len(table[0]) if table else 0
    print(f"  [SN-PDF] 감지: {year}년 {month}월, 테이블 {len(table)}행 x {ncols}열")

    menus = {}
    row_idx = 0

    while row_idx < len(table):
        row = table[row_idx]

        # 주차 행 감지: col[0]에 "N주차" 텍스트
        first = str(row[0] or '').strip()
        if not re.match(r'\d+주차', first):
            row_idx += 1
            continue

        # 이 행에서 날짜 추출 (각 열)
        date_cols = {}  # col_index → day_number
        for ci in ALL_DAY_COLS:
            if ci >= len(row):
                continue
            cell = str(row[ci] or '').strip()
            dn = _extract_date_from_header(cell)
            if dn and 1 <= dn <= 31:
                date_cols[ci] = dn

        if not date_cols:
            row_idx += 1
            continue

        # 다음 행들에서 메뉴 추출 (오전간식, 점심, 오후간식)
        sections = [
            (1, '오전간식'),
            (2, '점심'),
            (3, '오후간식'),
        ]

        for ci, dn in date_cols.items():
            try:
                ds = date(year, month, dn).isoformat()
            except ValueError:
                continue

            day_items = []
            for offset, sec_name in sections:
                ri = row_idx + offset
                if ri >= len(table):
                    break
                sec_row = table[ri]
                if ci >= len(sec_row):
                    continue
                cell_text = str(sec_row[ci] or '')
                for item in _parse_menu_cell(cell_text):
                    item['sec'] = sec_name
                    day_items.append(item)

            if day_items:
                menus[ds] = day_items

        # 다음 주 블록 (6행: 주차, 간식, 점심, 간식, 열량, 원산지)
        row_idx += 6

    print(f"  [SN-PDF] {len(menus)}일 추출, 총 {sum(len(v) for v in menus.values())}개 아이템")
    return {"year": year, "month": month, "menus": menus}


def parse_sn_excel(xlsx_path: str) -> dict:
    """
    성남시 레시피 Excel 파싱 (주별 시트).

    Returns:
        { "메뉴명": "조리방법 텍스트", ... }
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    recipes = {}
    # 주별 시트: "4월 1주(1일~4일)" 형태 또는 "생일식단"
    week_sheets = [s for s in wb.sheetnames if re.search(r'\d+주|\d+월', s)]

    for sheet_name in week_sheets:
        ws = wb[sheet_name]

        # 헤더 찾기
        header_row = None
        col_food = None
        col_recipe = None

        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or '').strip()
                if val == '음식명':
                    header_row = r
                    col_food = c
                elif val == '조리방법':
                    col_recipe = c
            if header_row and col_food and col_recipe:
                break

        if not header_row or not col_food or not col_recipe:
            continue

        # 데이터 행 읽기
        for r in range(header_row + 1, ws.max_row + 1):
            food_val = ws.cell(r, col_food).value
            recipe_val = ws.cell(r, col_recipe).value

            if not food_val:
                continue
            food_name = str(food_val).strip()
            # _x000D_ 아티팩트 제거
            food_name = food_name.replace('_x000D_', '').strip()
            # 연령별 접두어 제거
            food_name = re.sub(r'\(만\s*\d+-?\d*세\)\s*\n?', '', food_name).strip()
            # 장식 문자 제거
            food_name = DECO_CHARS.sub('', food_name).strip()
            # 공백만 남으면 스킵
            if not food_name:
                continue

            recipe_text = str(recipe_val).strip() if recipe_val else ''
            recipe_text = recipe_text.replace('_x000D_', '')

            if recipe_text and recipe_text != 'None' and recipe_text != '-':
                if food_name not in recipes:
                    recipes[food_name] = recipe_text

    print(f"  [SN-Excel] {len(recipes)}개 레시피 추출")
    return recipes


def _match_recipe(menu_name: str, recipes: dict) -> str:
    """메뉴명으로 레시피 매칭 (정확 → 공백무시 → 부분)"""
    if menu_name in recipes:
        return recipes[menu_name]
    # 공백 제거 후 매칭
    clean = menu_name.replace(' ', '')
    for rname, rtext in recipes.items():
        if rname.replace(' ', '') == clean:
            return rtext
    # 부분 매칭
    for rname, rtext in recipes.items():
        if rname in menu_name or menu_name in rname:
            return rtext
    return ''


def build_sn_json(pdf_data: dict, recipes: dict | None = None) -> dict:
    """성남시 파싱 결과를 표준 JSON 포맷으로 변환."""
    year = pdf_data['year']
    month = pdf_data['month']
    raw_menus = pdf_data['menus']
    recipes = recipes or {}

    matched = 0
    total = 0
    menu_data = {}
    for ds, items in sorted(raw_menus.items()):
        js_items = []
        for item in items:
            name = item['name']
            recipe = _match_recipe(name, recipes)
            if recipe:
                matched += 1
            total += 1
            js_items.append({
                "name": name,
                "recipe": recipe,
                "allergy": item.get('allergy', []),
                "sec": item.get('sec', '점심'),
            })
        menu_data[ds] = js_items

    if total > 0:
        print(f"  [SN-JSON] 레시피 매칭: {matched}/{total} ({matched*100//total}%)")

    # 주간 배열
    first_day = date(year, month, 1)
    days_since_monday = first_day.weekday()
    week_start = first_day - timedelta(days=days_since_monday)
    weeks = []
    d = week_start
    for _ in range(6):
        wk = []
        for _ in range(7):
            if d.weekday() != 6:  # 일요일 제외
                wk.append(d.isoformat() if d.month == month else None)
            d += timedelta(days=1)
        if any(wk):
            weeks.append(wk)

    return {
        "year": year,
        "month": month,
        "menus": menu_data,
        "allergens": {str(k): v for k, v in ALLERGEN_NAMES.items()},
        "weeks": weeks,
    }


# ── CLI ──
if __name__ == '__main__':
    import json
    import sys

    if len(sys.argv) < 2:
        print("Usage: python3 parse_seongnam.py <식단표.pdf> [레시피.xlsx] [output.json]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) >= 3 and sys.argv[2].endswith('.xlsx') else None
    out_path = sys.argv[-1] if sys.argv[-1].endswith('.json') else Path(pdf_path).stem + '.json'

    pdf_data = parse_sn_pdf(pdf_path)
    recipes = parse_sn_excel(xlsx_path) if xlsx_path else {}
    data = build_sn_json(pdf_data, recipes)

    Path(out_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n[OK] → {out_path}")
    print(f"  {data['year']}년 {data['month']}월, {len(data['menus'])}일, "
          f"{sum(len(v) for v in data['menus'].values())}개 아이템")
