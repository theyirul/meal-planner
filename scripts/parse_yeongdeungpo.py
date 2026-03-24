#!/usr/bin/env python3
"""
매일아침 — 영등포구 포맷 파서

영등포구어린이급식관리지원센터 식단표 PDF + 레시피 Excel 파서.
광진구 포맷과 다른 점:
  - PDF: 1페이지 대형 테이블 (33x19), 원형숫자(①②) 알레르기 표기
  - Excel: 주별 시트 (1주~5주), 요리법 포함

사용법:
  from parse_yeongdeungpo import parse_yd_pdf, parse_yd_excel, build_yd_json
"""

import re
from datetime import date, timedelta
from pathlib import Path

import pdfplumber

# ── 19종 알레르기 ──
ALLERGEN_NAMES = {
    1: "난류(계란)", 2: "우유", 3: "메밀", 4: "땅콩", 5: "대두",
    6: "밀", 7: "고등어", 8: "게", 9: "새우", 10: "돼지고기",
    11: "복숭아", 12: "토마토", 13: "아황산류", 14: "호두", 15: "닭고기",
    16: "쇠고기", 17: "오징어", 18: "조개류", 19: "잣"
}

# 원형숫자 → 일반숫자 매핑
CIRCLED_NUMS = {
    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10,
    '⑪': 11, '⑫': 12, '⑬': 13, '⑭': 14, '⑮': 15,
    '⑯': 16, '⑰': 17, '⑱': 18, '⑲': 19,
}
CIRCLED_PATTERN = re.compile('[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲]+')

# 요일 컬럼 인덱스 (19열 테이블)
DAY_COLS = [1, 5, 8, 11, 13, 16]  # 월 화 수 목 금 토

# 각 주 블록 내 행 오프셋
ROW_DATE = 0
ROW_AM_SNACK = 1
ROW_LUNCH = 2
ROW_PM_SNACK = 3
ROW_NUTRITION = 4

# 메뉴명에서 제거할 장식 문자
DECO_CHARS = re.compile(r'[♬★◉\U000f0074]')


def _extract_allergy_from_circled(text: str) -> tuple[str, list[int]]:
    """
    원형숫자가 포함된 텍스트에서 메뉴명과 알레르기 번호를 분리.
    "쇠고기장조림⑤⑥⑯" → ("쇠고기장조림", [5, 6, 16])
    """
    nums = []
    clean = text
    for match in CIRCLED_PATTERN.finditer(text):
        for ch in match.group():
            if ch in CIRCLED_NUMS:
                nums.append(CIRCLED_NUMS[ch])
        clean = clean[:match.start()] + clean[match.end():]
    # 괄호형 알레르기도 처리 (혹시 섞여있을 때)
    paren_pat = re.compile(r'[\(（]([\d,.\s]+)[\)）]')
    for m in paren_pat.finditer(clean):
        for n in re.findall(r'\d+', m.group(1)):
            v = int(n)
            if 1 <= v <= 19 and v not in nums:
                nums.append(v)
        clean = clean[:m.start()] + clean[m.end():]
    clean = clean.strip()
    return clean, sorted(set(nums))


def _parse_menu_cell(cell_text: str) -> list[dict]:
    """
    한 셀의 텍스트에서 개별 메뉴 아이템을 추출.
    줄바꿈으로 구분, /로 구분, 각 아이템에서 원형숫자 알레르기 분리.
    """
    if not cell_text or not cell_text.strip():
        return []

    items = []
    # 괄호 안 보조설명을 메인 이름에 합치기
    # "오곡라떼②⑤\n(미숫가루⑤/우유②)" → "오곡라떼②⑤"로만 처리
    cell_text = re.sub(r'\n?\(([^)]*[/][^)]*)\)', '', cell_text)
    # "[마시는요구르트]②" → "마시는요구르트②"
    cell_text = re.sub(r'\[([^\]]+)\]', r'\1', cell_text)

    # 줄바꿈으로 1차 분리
    lines = cell_text.strip().split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # /로 구분된 아이템 분리 (예: "토마토스틱⑫\n/치즈②")
        parts = line.split('/')
        for part in parts:
            part = part.strip()
            if not part:
                continue
            # 장식 문자 제거
            part = DECO_CHARS.sub('', part).strip()
            # 열량/영양 정보 스킵
            if re.match(r'^\d+/\d+$', part):
                continue
            # 라벨 스킵
            if part in ('(공통)', '점심', '오전간식', '오후간식'):
                continue
            # "후식:" 등 접두어 제거
            part = re.sub(r'^후식\s*:\s*', '', part)
            if not part:
                continue

            name, allergy = _extract_allergy_from_circled(part)
            # 연령별 표기 "(만1-2세)" 등 제거
            name = re.sub(r'\(만\d+-?\d*세\)', '', name).strip()
            if name:
                items.append({"name": name, "allergy": allergy})

    return items


def _extract_date_num(cell_text: str) -> int | None:
    """날짜 셀에서 숫자 추출. "9", "10 [저 염] 국없데이", "2 [대체공휴일]" 등"""
    if not cell_text:
        return None
    m = re.match(r'(\d{1,2})', cell_text.strip())
    return int(m.group(1)) if m else None


def _find_main_table(tables: list) -> list | None:
    """PDF 테이블들 중 메인 식단 테이블 찾기 (가장 큰 테이블)"""
    if not tables:
        return None
    # 행 수가 가장 많은 테이블 선택
    return max(tables, key=lambda t: len(t))


def _detect_year_month_from_pdf(pdf) -> tuple[int, int]:
    """PDF 전체 텍스트에서 연도/월 감지"""
    text = pdf.pages[0].extract_text() or ''

    # "(3월 식단)" 패턴 — 영등포구 특유
    m = re.search(r'\((\d{1,2})월\s*식단\)', text)
    if m:
        month = int(m.group(1))
        # 연도는 발행일에서 추출
        ym = re.search(r'(\d{4})\.\s*\d{1,2}\.\s*\d{1,2}', text)
        year = int(ym.group(1)) if ym else date.today().year
        return year, month

    # "2026년 3월" 패턴
    m = re.search(r'(\d{4})\s*[년.]\s*(\d{1,2})\s*월', text)
    if m:
        return int(m.group(1)), int(m.group(2))

    # 발행일에서 추론 (다음 달)
    m = re.search(r'(\d{4})\.\s*(\d{1,2})\.\s*\d{1,2}', text)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        return (y, mo + 1) if mo < 12 else (y + 1, 1)

    raise ValueError("연도/월을 감지할 수 없습니다")


def parse_yd_pdf(pdf_path: str) -> dict:
    """
    영등포구 식단표 PDF 파싱.

    Returns:
        {
            "year": 2026, "month": 3,
            "menus": {
                "2026-03-03": [
                    {"name": "토마토스틱", "allergy": [12], "sec": "오전간식"},
                    ...
                ], ...
            }
        }
    """
    with pdfplumber.open(pdf_path) as pdf:
        year, month = _detect_year_month_from_pdf(pdf)
        tables = pdf.pages[0].extract_tables()

    table = _find_main_table(tables)
    if not table or len(table) < 10:
        raise ValueError(f"메인 테이블을 찾을 수 없습니다 ({len(table) if table else 0} rows)")

    print(f"  [YD-PDF] 감지: {year}년 {month}월, 테이블 {len(table)}행 x {len(table[0])}열")

    menus = {}

    # 주 블록 시작 행 찾기: col[0]이 비어있고, 숫자가 있는 행이 날짜 행
    row_idx = 0
    while row_idx < len(table):
        row = table[row_idx]

        # 날짜 행 감지: col[0]이 비어있거나 공백, 그리고 DAY_COLS 중 하나에 숫자가 있음
        first_col = str(row[0] or '').strip()
        if first_col and first_col not in ('', ' '):
            # 오전간식/점심/오후간식/열량 등 데이터 행이면 스킵
            row_idx += 1
            continue

        # DAY_COLS에서 날짜 숫자 찾기
        found_dates = False
        for ci in DAY_COLS:
            if ci < len(row):
                dn = _extract_date_num(str(row[ci] or ''))
                if dn and 1 <= dn <= 31:
                    found_dates = True
                    break

        if not found_dates:
            row_idx += 1
            continue

        # 이 행이 날짜 행. 다음 행들에서 메뉴 추출
        date_row = row
        # 남은 행이 충분한지 확인
        remaining = len(table) - row_idx
        has_am = remaining > ROW_AM_SNACK
        has_lunch = remaining > ROW_LUNCH
        has_pm = remaining > ROW_PM_SNACK

        for ci in DAY_COLS:
            if ci >= len(date_row):
                continue
            dn = _extract_date_num(str(date_row[ci] or ''))
            if not dn:
                continue
            try:
                ds = date(year, month, dn).isoformat()
            except ValueError:
                continue

            day_items = []

            # 오전간식
            if has_am:
                am_row = table[row_idx + ROW_AM_SNACK]
                am_cell = str(am_row[ci] or '') if ci < len(am_row) else ''
                for item in _parse_menu_cell(am_cell):
                    item['sec'] = '오전간식'
                    day_items.append(item)

            # 점심
            if has_lunch:
                lunch_row = table[row_idx + ROW_LUNCH]
                lunch_cell = str(lunch_row[ci] or '') if ci < len(lunch_row) else ''
                for item in _parse_menu_cell(lunch_cell):
                    item['sec'] = '점심'
                    day_items.append(item)

            # 오후간식
            if has_pm:
                pm_row = table[row_idx + ROW_PM_SNACK]
                pm_cell = str(pm_row[ci] or '') if ci < len(pm_row) else ''
                for item in _parse_menu_cell(pm_cell):
                    item['sec'] = '오후간식'
                    day_items.append(item)

            if day_items:
                menus[ds] = day_items

        # 다음 주 블록으로 (날짜+4행 = 5행 블록)
        row_idx += 5

    print(f"  [YD-PDF] {len(menus)}일 추출, 총 {sum(len(v) for v in menus.values())}개 아이템")
    return {"year": year, "month": month, "menus": menus}


def parse_yd_excel(xlsx_path: str) -> dict:
    """
    영등포구 레시피 Excel 파싱 (주별 시트 1주~5주).

    Returns:
        { "메뉴명": "조리방법 텍스트", ... }
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    recipes = {}
    week_sheets = [s for s in wb.sheetnames if re.match(r'\d+주', s)]

    for sheet_name in week_sheets:
        ws = wb[sheet_name]
        # 헤더 찾기: "음식명" 컬럼과 "조리방법" 컬럼
        header_row = None
        col_food = None
        col_recipe = None

        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or '').strip()
                if val == '음식명':
                    header_row = r
                    col_food = c
                elif val == '조리방법':
                    col_recipe = c
            if header_row:
                break

        if not header_row or not col_food or not col_recipe:
            continue

        # 데이터 행 읽기
        for r in range(header_row + 2, ws.max_row + 1):  # +2: 서브헤더 스킵
            food_val = ws.cell(r, col_food).value
            recipe_val = ws.cell(r, col_recipe).value

            if not food_val:
                continue
            food_name = str(food_val).strip()
            # 연령별 접두어 제거: "(만1-2세)\n흑미밥", "(만 1-2세)\n모닝빵"
            food_name = re.sub(r'\(만\s*\d+-?\d*세\)\s*\n?', '', food_name).strip()
            # _x000D_ 아티팩트 제거
            food_name = food_name.replace('_x000D_', '').strip()
            # 장식 문자 제거
            food_name = DECO_CHARS.sub('', food_name).strip()

            if not food_name:
                continue

            recipe_text = str(recipe_val).strip() if recipe_val else ''
            # _x000D_ 제거 (Excel 줄바꿈 아티팩트)
            recipe_text = recipe_text.replace('_x000D_', '')

            if recipe_text and food_name not in recipes:
                recipes[food_name] = recipe_text

    print(f"  [YD-Excel] {len(recipes)}개 레시피 추출")
    return recipes


def _match_recipe(menu_name: str, recipes: dict) -> str:
    """메뉴명으로 레시피 매칭 (정확 매칭 → 부분 매칭)"""
    # 정확 매칭
    if menu_name in recipes:
        return recipes[menu_name]
    # 장식 제거 후 매칭
    clean = DECO_CHARS.sub('', menu_name).strip()
    if clean in recipes:
        return recipes[clean]
    # 부분 매칭: 레시피 키가 메뉴명에 포함되거나 반대
    for rname, rtext in recipes.items():
        if rname in menu_name or menu_name in rname:
            return rtext
    return ''


def build_yd_json(pdf_data: dict, recipes: dict | None = None) -> dict:
    """
    영등포구 파싱 결과를 build.py/index.html JSON 포맷으로 변환.
    """
    year = pdf_data['year']
    month = pdf_data['month']
    raw_menus = pdf_data['menus']
    recipes = recipes or {}

    menu_data = {}
    for ds, items in sorted(raw_menus.items()):
        js_items = []
        for item in items:
            name = item['name']
            js_item = {
                "name": name,
                "recipe": _match_recipe(name, recipes),
                "allergy": item.get('allergy', []),
                "sec": item.get('sec', '점심'),
            }
            js_items.append(js_item)
        menu_data[ds] = js_items

    # 주간 배열
    first_day = date(year, month, 1)
    days_since_monday = first_day.weekday()
    week_start = first_day - timedelta(days=days_since_monday)
    weeks = []
    d = week_start
    for _ in range(6):
        wk = []
        for _ in range(7):
            if d.weekday() != 6:
                wk.append(d.isoformat() if d.month == month else None)
            d += timedelta(days=1)
        if any(wk):
            weeks.append(wk)

    return {
        "year": year,
        "month": month,
        "menus": menu_data,
        "allergens": {str(k): v for k, v in ALLERGEN_NAMES.items()},
        "weeks": weeks,
    }


# ── CLI ──
if __name__ == '__main__':
    import json
    import sys

    if len(sys.argv) < 2:
        print("Usage: python3 parse_yeongdeungpo.py <식단표.pdf> [레시피.xlsx] [output.json]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) >= 3 and sys.argv[2].endswith('.xlsx') else None
    out_path = sys.argv[-1] if sys.argv[-1].endswith('.json') else Path(pdf_path).stem + '.json'

    pdf_data = parse_yd_pdf(pdf_path)
    recipes = parse_yd_excel(xlsx_path) if xlsx_path else {}
    data = build_yd_json(pdf_data, recipes)

    Path(out_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"\n[OK] → {out_path}")
    print(f"  {data['year']}년 {data['month']}월, {len(data['menus'])}일, "
          f"{sum(len(v) for v in data['menus'].values())}개 아이템")
