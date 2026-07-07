#!/usr/bin/env python3
"""식단표 ↔ 조리지시서(xlsx) 교차검증.

조리지시서 xlsx가 있을 때 build.py가 자동 호출한다(룰 지침: 있으면 항상 함께 검증).
두 축으로 대조한다:
  1) 배정 검증 — '에너지 및 영양소 분석표' 시트의 (요일·끼니·음식명)과 파싱 메뉴를 비교.
     날짜/끼니 밀림·누락·오타를 잡는다.
  2) 알레르기 누락 flag — '조리지시서(N주차)' 시트의 식재료로 알레르기를 추정,
     식단표 표기에 빠진 후보를 알린다(휴리스틱 → 참고용, 원본이 미표기면 그대로 둘 수 있음).

의존성: openpyxl.
"""
from __future__ import annotations
import re

CIRCLED_RE = re.compile('[' + ''.join(chr(0x2460 + i) for i in range(20)) + ']')

# 식재료 키워드 → 알레르기 번호 (보수적. 오탐 줄이려 부분일치 주의)
ALLERGEN_ING = {
    1: ['달걀', '계란', '메추리알', '난백', '난황', '마요'],
    2: ['우유', '치즈', '버터', '생크림', '크림', '분유', '연유', '요구르트', '요거트'],
    3: ['메밀'], 4: ['땅콩'],
    5: ['대두', '두부', '된장', '간장', '두유', '유부', '청국장', '콩나물', '콩가루', '쌈장', '춘장', '고추장'],
    6: ['밀가루', '국수', '우동', '스파게티', '파스타', '부침가루', '튀김가루', '빵가루', '카레', '당면', '수제비', '만두피', '부침', '핫도그', '베이글', '빵'],
    7: ['고등어'], 8: ['게맛살', '크래미', '꽃게'], 9: ['새우'],
    10: ['돼지', '돈육', '돈까스', '돈가스', '베이컨', '햄', '소시지', '폭찹', '제육'],
    11: ['복숭아'], 12: ['토마토', '케첩', '케찹'], 14: ['호두'],
    15: ['닭', '계육'], 16: ['소고기', '쇠고기', '우육', '한우', '차돌', '불고기'],
    17: ['오징어'], 18: ['조개', '굴소스', '전복', '홍합', '바지락', '재첩', '꼬막'], 19: ['잣'],
}
ALLERGEN_NAME = {1: '난류', 2: '우유', 3: '메밀', 4: '땅콩', 5: '대두', 6: '밀', 7: '고등어',
                 8: '게', 9: '새우', 10: '돼지', 11: '복숭아', 12: '토마토', 13: '아황산',
                 14: '호두', 15: '닭', 16: '소고기', 17: '오징어', 18: '조개', 19: '잣'}

SECS = {'오전간식', '점심', '오후간식'}


def _norm(s):
    s = CIRCLED_RE.sub('', str(s)).replace('♠', '')
    s = re.sub(r'[\(\[][^)\]]*[\)\]]', '', s)   # 괄호/대괄호 내용 제거
    return re.sub(r'\s+', '', s).strip()


def _implied(ings):
    out = {}
    for ing in ings:
        for num, kws in ALLERGEN_ING.items():
            if any(k in ing for k in kws):
                out.setdefault(num, []).append(ing)
    return out


def crosscheck(xlsx_path: str, menus: dict) -> dict | None:
    """menus={'YYYY-MM-DD':[{name,allergy,sec}]}. 조리지시서 없으면 None."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return None

    nutri = next((s for s in wb.sheetnames if '영양소 분석표' in s and '테마' not in s), None)
    if not nutri:
        return None  # 조리지시서 형식 아님

    # 1) 배정 검증용: (day,sec)->set(norm 음식명)
    ws = wb[nutri]
    xls = {}
    day = sec = None
    for r in range(6, ws.max_row + 1):
        a, b, c = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if a:
            m = re.match(r'(\d{1,2})', str(a).strip())
            if m:
                day = int(m.group(1))
        if b and str(b).strip() in SECS:
            sec = str(b).strip()
        if c and day and sec:
            nm = _norm(c)
            if nm and nm != '음식명' and not nm.startswith('*'):
                xls.setdefault((day, sec), set()).add(nm)

    hwp = {}
    for dt, items in menus.items():
        d = int(dt[-2:])
        for it in items:
            hwp.setdefault((d, it['sec']), set()).add(_norm(it['name']))

    mis = []  # 배정 불일치 (표기변형 제외 어려우니 전량 보고, 판단은 사람)
    for k in sorted(set(xls) | set(hwp)):
        ox, oh = xls.get(k, set()) - hwp.get(k, set()), hwp.get(k, set()) - xls.get(k, set())
        if ox or oh:
            mis.append((k, sorted(ox), sorted(oh)))

    # 2) 알레르기 누락 flag: 주차별 조리지시서 식재료
    recipe_ings = {}
    for sn in wb.sheetnames:
        if not sn.startswith('조리지시서'):
            continue
        w = wb[sn]
        dish = None
        for r in range(7, w.max_row + 1):
            nm, ing = w.cell(r, 3).value, w.cell(r, 4).value
            if nm and str(nm).strip():
                dish = _norm(nm)
            if dish and ing:
                recipe_ings.setdefault(dish, set()).add(str(ing).strip())

    gaps = []
    for dt, items in sorted(menus.items()):
        for it in items:
            ings = recipe_ings.get(_norm(it['name']))
            if not ings:
                continue
            imp = _implied(ings)
            miss = set(imp) - set(it['allergy'])
            if miss:
                gaps.append((dt, it['sec'], it['name'], sorted(it['allergy']),
                             {ALLERGEN_NAME.get(n, n): imp[n] for n in sorted(miss)}))

    return {'assign_mismatch': mis, 'allergen_gaps': gaps, 'blocks': len(set(xls) | set(hwp))}


def print_report(report: dict, region: str = ''):
    if not report:
        return
    print(f"\n  ── 교차검증{(' [' + region + ']') if region else ''} ──")
    mis = report['assign_mismatch']
    print(f"  배정 대조: {report['blocks']}블록 중 불일치 {len(mis)}건 (표기변형 포함, 사람 확인)")
    for (day, sec), ox, oh in mis:
        d = []
        if ox: d.append(f"지시서만={ox}")
        if oh: d.append(f"식단표만={oh}")
        print(f"    {day}일 {sec}: {' / '.join(d)}")
    gaps = report['allergen_gaps']
    print(f"  알레르기 누락 후보: {len(gaps)}건 (식재료엔 있으나 식단표 미표기 — 원본 미표기면 유지 가능)")
    for dt, sec, nm, have, hits in gaps:
        srcs = '; '.join(f"+{al}?({','.join(v[:2])})" for al, v in hits.items())
        print(f"    {dt} {sec} {nm} 표기{have} → {srcs}")
